"""
St. Anne Mission Hospital — ICT Command Centre v2.0
Local browser app + HTML UI + Excel as live database
Enhancements: Custom departments/types, item transfers, repairs, replacements,
              classified role access, auto-audit logging, portable app builder
"""
import os as _os
_os.environ.setdefault("PYWEBVIEW_GUI", "edgechromium")
_os.environ.setdefault("WEBKIT_DISABLE_COMPOSITING_MODE", "1")
_os.environ.setdefault("PYWEBVIEW_NO_UIAUTOMATION", "1")

import json, os, sys, datetime, shutil, openpyxl, threading, re, hashlib, secrets
import http.server, socketserver, webbrowser, urllib.parse
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
MASTER_XL  = os.path.join(BASE_DIR, "ICT_MASTER.xlsx")
BACKUP_DIR = os.path.join(BASE_DIR, "backups")
AUTH_FILE  = os.path.join(BASE_DIR, "auth.json")
CONFIG_FILE= os.path.join(BASE_DIR, "config.json")

SHEET_INV     = "Inventory"
SHEET_HISTORY = "History"
SHEET_REPLACE = "Replacements"
SHEET_INK     = "Ink"
SHEET_LOG     = "Audit Log"
SHEET_CONFIG  = "Config"
ALL_SHEETS    = [SHEET_INV, SHEET_HISTORY, SHEET_REPLACE, SHEET_INK, SHEET_LOG, SHEET_CONFIG]

INV_COLS = [
    "Date Collected","Purchase Date","Equipment Type","Brand/Model","Serial No.",
    "Assigned To","Department/Location","Condition/Status","OS/Firmware",
    "IP Address","MAC Address","CPU","RAM","Disk","Remarks","Transfer History"
]
INK_CODES    = ["BK","C","LC","M","LM","Y"]
INK_NAMES    = {"BK":"Black","C":"Cyan","LC":"Light Cyan","M":"Magenta","LM":"Light Magenta","Y":"Yellow"}
INK_DEFAULTS = {"BK":9,"C":3,"LC":5,"M":1,"LM":5,"Y":11}

# Default lists — users can extend these
DEFAULT_DEPTS = [
    "ICT Department","Administration","Outpatient","Inpatient","Pharmacy",
    "Laboratory","Radiology","Theatre","Maternity","Accounts","Records","Store"
]
DEFAULT_TYPES = [
    "Desktop Computer","Laptop","Printer","Scanner","Server","Switch","Router",
    "UPS","Monitor","Keyboard","Mouse","External HDD","Flash Drive","Projector",
    "Tablet","Phone","Photocopier","Camera","Other"
]
DEFAULT_CONDITIONS = [
    "New","Good","Fair","Fair (Needs Repair)","Needs Repair","Replaced","Disposed"
]

def today(): return datetime.date.today().strftime("%Y-%m-%d")
def now():   return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# ════════════════════════════════════════════
# CONFIG ENGINE
# ════════════════════════════════════════════
class ConfigEngine:
    def __init__(self):
        self._lock = threading.Lock()
        self._data = self._load()

    def _load(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE) as f:
                    return json.load(f)
            except: pass
        return {"departments": DEFAULT_DEPTS[:], "types": DEFAULT_TYPES[:], "conditions": DEFAULT_CONDITIONS[:]}

    def _save(self):
        with open(CONFIG_FILE, 'w') as f:
            json.dump(self._data, f, indent=2)

    def get(self):
        with self._lock:
            return dict(self._data)

    def add_item(self, key, value):
        with self._lock:
            v = value.strip()
            if v and v not in self._data.get(key, []):
                self._data.setdefault(key, []).append(v)
                self._save()
                return True
            return False

    def remove_item(self, key, value):
        with self._lock:
            lst = self._data.get(key, [])
            if value in lst:
                lst.remove(value)
                self._save()
                return True
            return False

# ════════════════════════════════════════════
# EXCEL ENGINE
# ════════════════════════════════════════════
class ExcelEngine:
    def __init__(self):
        self._lock = threading.Lock()
        self._ensure_workbook()

    def _init_headers(self, ws, name):
        hdrs = {
            SHEET_INV:     INV_COLS,
            SHEET_HISTORY: ["Date","Asset","Serial No.","From Department","To Department","Event","Description","By"],
            SHEET_REPLACE: ["Date Replaced","Old Asset","Old Serial","Department","Replaced With","New Serial","New Condition","Note","By"],
            SHEET_INK:     ["Timestamp","Colour Code","Colour Name","Action","Quantity","Big ICT","Small ICT","Note"],
            SHEET_LOG:     ["Timestamp","Action","Description","By"],
            SHEET_CONFIG:  ["Key","Value"],
        }
        ws.append(hdrs.get(name, []))
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", fgColor="1F2937")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

    def _set_widths(self, ws):
        widths = {
            "Date Collected":13,"Purchase Date":13,"Equipment Type":22,"Brand/Model":20,
            "Serial No.":22,"Assigned To":20,"Department/Location":24,"Condition/Status":22,
            "OS/Firmware":16,"IP Address":16,"MAC Address":18,"CPU":14,"RAM":12,"Disk":20,
            "Remarks":32,"Transfer History":40
        }
        if ws.max_row < 1: return
        h = [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        for i, col_name in enumerate(h, 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(col_name, 16)

    def _write_log(self, wb, action, desc, by="ICT Manager"):
        wb[SHEET_LOG].append([now(), action, desc, by])

    def _backup(self):
        os.makedirs(BACKUP_DIR, exist_ok=True)
        ts  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        dst = os.path.join(BACKUP_DIR, f"ICT_backup_{ts}.xlsx")
        shutil.copy2(MASTER_XL, dst)

    def _ensure_workbook(self):
        if not os.path.exists(MASTER_XL):
            self._create_fresh()
            return
        wb = openpyxl.load_workbook(MASTER_XL)
        OLD_DATA_SHEETS = ["work sheet","Work Sheet","worksheet","WorkSheet"]
        old_data_sheet  = next((s for s in OLD_DATA_SHEETS if s in wb.sheetnames), None)
        needs_migration = SHEET_INV not in wb.sheetnames and old_data_sheet is not None
        if needs_migration:
            os.makedirs(BACKUP_DIR, exist_ok=True)
            shutil.copy2(MASTER_XL, os.path.join(BACKUP_DIR, "ICT_original_before_migration.xlsx"))
            old_rows = self._read_old_rows(wb, old_data_sheet)
            for sname in list(wb.sheetnames):
                del wb[sname]
            for sheet in ALL_SHEETS:
                ws = wb.create_sheet(sheet)
                self._init_headers(ws, sheet)
            ws_inv = wb[SHEET_INV]
            for row in old_rows:
                # pad to new column count
                while len(row) < len(INV_COLS):
                    row.append('')
                ws_inv.append(row)
            self._set_widths(ws_inv)
            wb.save(MASTER_XL); wb.close()
            return
        changed = False
        for sheet in ALL_SHEETS:
            if sheet not in wb.sheetnames:
                ws = wb.create_sheet(sheet)
                self._init_headers(ws, sheet)
                changed = True
        # Add Transfer History column if missing
        if SHEET_INV in wb.sheetnames:
            ws = wb[SHEET_INV]
            if ws.max_row >= 1:
                h = [str(c.value).strip() if c.value else '' for c in ws[1]]
                if "Transfer History" not in h:
                    col = len(h) + 1
                    ws.cell(1, col).value = "Transfer History"
                    ws.cell(1, col).font  = Font(bold=True, color="FFFFFF", size=11)
                    ws.cell(1, col).fill  = PatternFill("solid", fgColor="1F2937")
                    changed = True
        if changed:
            wb.save(MASTER_XL)
        wb.close()

    def _read_old_rows(self, wb, sheet_name):
        ws   = wb[sheet_name]
        rows = list(ws.values)
        if not rows: return []
        old_h = [str(c).strip() if c else '' for c in rows[0]]
        def g(row, name, default=''):
            try:
                i = old_h.index(name)
                v = row[i] if i < len(row) else None
                if v is None: return default
                if isinstance(v, (datetime.date, datetime.datetime)):
                    return v.strftime('%Y-%m-%d')
                s = str(v).strip()
                if re.match(r'\d{4}-\d{2}-\d{2}[ T]', s): s = s[:10]
                return default if s.lower() in ['nan','none','nat','<na>'] else s
            except ValueError: return default
        def fix_ip(ip): return (ip or '').replace('192.068.','192.168.').replace('192.0.168.','192.168.')
        result = []
        for row in rows[1:]:
            if not any(v for v in row): continue
            ram  = ' '.join(p for p in [g(row,'RAM Size'),g(row,'RAM Type')] if p)
            disk = ' '.join(p for p in [g(row,'Disk Size'),g(row,'Disk Type'),g(row,'Disk Health')] if p)
            result.append([
                g(row,'Date Collected'), g(row,'Purchase Date'), g(row,'Equipment Type'),
                g(row,'Brand/Model'), g(row,'Serial No.'), g(row,'Assigned To'),
                g(row,'Department/Location').strip(), g(row,'Condition/Status') or 'Good',
                g(row,'OS/Firmware'), fix_ip(g(row,'IP Address')), g(row,'MAC Address'),
                g(row,'CPU'), ram, disk, g(row,'Remarks'), ''
            ])
        return result

    def _create_fresh(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for s in ALL_SHEETS:
            self._init_headers(wb.create_sheet(s), s)
        wb.save(MASTER_XL)

    def _clean(self, v):
        if v is None: return ''
        if isinstance(v, (datetime.date, datetime.datetime)): return v.strftime('%Y-%m-%d')
        s = str(v).strip()
        if re.match(r'\d{4}-\d{2}-\d{2}[ T]', s): s = s[:10]
        return '' if s.lower() in ['nan','none','nat','<na>'] else s

    # ── READ ─────────────────────────────────
    def load_all(self):
        with self._lock:
            wb   = openpyxl.load_workbook(MASTER_XL, data_only=True)
            inv  = self._read_inv(wb)
            ink  = self._read_ink(wb)
            log  = self._read_log(wb)
            hist = self._read_hist(wb)
            wb.close()
            return {"inventory": inv, "ink": ink, "log": log, "history": hist}

    def _read_inv(self, wb):
        ws   = wb[SHEET_INV]
        rows = list(ws.values)
        if len(rows) < 2: return []
        h   = [str(c).strip() if c else '' for c in rows[0]]
        out = []
        for i, row in enumerate(rows[1:], start=2):
            d = {'_row': i}
            for j, col in enumerate(h):
                d[col] = self._clean(row[j] if j < len(row) else None)
            out.append(d)
        return out

    def _read_ink(self, wb):
        ws    = wb[SHEET_INK]
        rows  = list(ws.values)
        state = {k: {"store": INK_DEFAULTS[k], "bigIct": 1, "smallIct": 1} for k in INK_CODES}
        for row in rows[1:]:
            if not row or not row[1]: continue
            code   = str(row[1]).strip()
            action = str(row[3]).strip() if len(row)>3 and row[3] else ''
            try: qty = int(float(str(row[4]))) if len(row)>4 and row[4] else 0
            except: qty = 0
            try: big = int(float(str(row[5]))) if len(row)>5 and row[5] else None
            except: big = None
            try: sml = int(float(str(row[6]))) if len(row)>6 and row[6] else None
            except: sml = None
            if code in state:
                if action == 'restock': state[code]['store'] += qty
                elif action == 'use':   state[code]['store'] = max(0, state[code]['store'] - qty)
                if big is not None: state[code]['bigIct']   = big
                if sml is not None: state[code]['smallIct'] = sml
        return state

    def _read_log(self, wb):
        ws  = wb[SHEET_LOG]
        out = []
        for row in list(ws.values)[1:]:
            out.append({'time':str(row[0]) if row[0] else '','action':str(row[1]) if len(row)>1 and row[1] else '','desc':str(row[2]) if len(row)>2 and row[2] else ''})
        return list(reversed(out))

    def _read_hist(self, wb):
        ws  = wb[SHEET_HISTORY]
        out = []
        for row in list(ws.values)[1:]:
            out.append({
                'date':  str(row[0]) if row[0] else '',
                'asset': str(row[1]) if len(row)>1 and row[1] else '',
                'serial':str(row[2]) if len(row)>2 and row[2] else '',
                'from':  str(row[3]) if len(row)>3 and row[3] else '',
                'to':    str(row[4]) if len(row)>4 and row[4] else '',
                'event': str(row[5]) if len(row)>5 and row[5] else '',
                'note':  str(row[6]) if len(row)>6 and row[6] else '',
                'by':    str(row[7]) if len(row)>7 and row[7] else '',
            })
        return list(reversed(out))

    # ── WRITE ────────────────────────────────
    def add_asset(self, data):
        with self._lock:
            self._backup()
            wb = openpyxl.load_workbook(MASTER_XL)
            ws = wb[SHEET_INV]
            row_data = [data.get(c, '') for c in INV_COLS]
            ws.append(row_data)
            new_row = ws.max_row
            self._set_widths(ws)
            self._write_log(wb, 'ADD', f"Added: {data.get('Equipment Type','')} — {data.get('Brand/Model','')} | {data.get('Department/Location','')}")
            wb.save(MASTER_XL); wb.close()
            return new_row

    def update_asset(self, row_num, data):
        with self._lock:
            self._backup()
            wb = openpyxl.load_workbook(MASTER_XL)
            ws = wb[SHEET_INV]
            h  = [ws.cell(1,c).value for c in range(1,len(INV_COLS)+2)]
            for j, col in enumerate(h, 1):
                if col in data:
                    ws.cell(row=row_num, column=j).value = data[col]
            self._write_log(wb, 'EDIT', f"Edited row {row_num}: {data.get('Equipment Type','')} — {data.get('Brand/Model','')}")
            wb.save(MASTER_XL); wb.close()

    def delete_asset(self, row_num, desc):
        with self._lock:
            self._backup()
            wb = openpyxl.load_workbook(MASTER_XL)
            wb[SHEET_INV].delete_rows(row_num)
            self._write_log(wb, 'DELETE', f"Deleted: {desc}")
            wb.save(MASTER_XL); wb.close()

    def transfer_asset(self, row_num, asset_desc, serial, from_dept, to_dept, note, date, by='ICT Manager'):
        """Transfer an asset to another department — updates Inventory + logs history."""
        with self._lock:
            self._backup()
            wb     = openpyxl.load_workbook(MASTER_XL)
            ws_inv = wb[SHEET_INV]
            h      = [ws_inv.cell(1,c).value for c in range(1, ws_inv.max_column+1)]
            def col_idx(name):
                try: return h.index(name)+1
                except: return None
            dept_col = col_idx("Department/Location")
            xfer_col = col_idx("Transfer History")
            if dept_col:
                ws_inv.cell(row=row_num, column=dept_col).value = to_dept
            if xfer_col:
                old_xfer = ws_inv.cell(row=row_num, column=xfer_col).value or ''
                entry = f"[{date}] {from_dept} → {to_dept}"
                if note: entry += f": {note}"
                ws_inv.cell(row=row_num, column=xfer_col).value = (str(old_xfer) + " | " + entry).strip(" |")
            # History
            by_val = by if by and by.strip() else "ICT Manager"
            wb[SHEET_HISTORY].append([date, asset_desc, serial, from_dept, to_dept, "transferred",
                f"Transferred from {from_dept} to {to_dept}. Note: {note}", by_val])
            self._write_log(wb, 'TRANSFER', f"Transferred {asset_desc} (S/N:{serial}) from {from_dept} → {to_dept}")
            wb.save(MASTER_XL); wb.close()

    def record_replacement(self, old_row, old_desc, rep):
        with self._lock:
            self._backup()
            wb     = openpyxl.load_workbook(MASTER_XL)
            ws_inv = wb[SHEET_INV]
            h      = [ws_inv.cell(1,c).value for c in range(1, len(INV_COLS)+2)]
            def col_idx(name):
                try: return h.index(name)+1
                except: return None
            ci = col_idx("Condition/Status"); ri = col_idx("Remarks")
            if ci: ws_inv.cell(row=old_row, column=ci).value = "Replaced"
            if ri:
                old_rem = ws_inv.cell(row=old_row, column=ri).value or ''
                ws_inv.cell(row=old_row, column=ri).value = (str(old_rem)+f" | REPLACED {rep['date']}: {rep['note']}").strip(" |")
            ws_inv.append([{
                "Date Collected":rep['date'],"Purchase Date":rep['date'],
                "Equipment Type":rep.get('newType',''),"Brand/Model":rep.get('newBrand',''),
                "Serial No.":rep.get('newSerial',''),"Assigned To":rep.get('assigned',''),
                "Department/Location":rep.get('dept',''),"Condition/Status":rep.get('newCond','New'),
                "OS/Firmware":"","IP Address":rep.get('ip',''),"MAC Address":"","CPU":"","RAM":"","Disk":"",
                "Remarks":f"Replaced: {old_desc} on {rep['date']}. {rep['note']}","Transfer History":""
            }.get(c,'') for c in INV_COLS])
            self._set_widths(ws_inv)
            wb[SHEET_REPLACE].append([rep['date'],old_desc,rep.get('oldSerial',''),rep.get('dept',''),
                f"{rep.get('newType','')} {rep.get('newBrand','')}".strip(),rep.get('newSerial',''),
                rep.get('newCond','New'),rep['note'],"ICT Manager"])
            wb[SHEET_HISTORY].append([rep['date'],old_desc,rep.get('oldSerial',''),rep.get('dept',''),
                rep.get('dept',''),"replaced",
                f"Replaced with: {rep.get('newType','')} {rep.get('newBrand','')} S/N:{rep.get('newSerial','')}. Note: {rep['note']}","ICT Manager"])
            self._write_log(wb,'REPLACE',f"Replaced {old_desc} → {rep.get('newBrand','')} on {rep['date']}")
            wb.save(MASTER_XL); wb.close()

    def write_history_event(self, asset_desc, serial, dept, event, note, date, cost='', by='ICT Manager'):
        with self._lock:
            wb = openpyxl.load_workbook(MASTER_XL)
            by_val = by if by and by.strip() else "ICT Manager"
            wb[SHEET_HISTORY].append([date, asset_desc, serial, dept, dept, event, note, by_val])
            if event == 'repaired':
                ws = wb[SHEET_INV]
                h  = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
                try: ser_col  = h.index("Serial No.")+1
                except: ser_col = None
                try: cond_col = h.index("Condition/Status")+1
                except: cond_col = None
                if ser_col and cond_col:
                    for row in range(2, ws.max_row+1):
                        if str(ws.cell(row,ser_col).value or '').strip() == serial.strip():
                            if ws.cell(row,cond_col).value in ["Fair (Needs Repair)","Needs Repair"]:
                                ws.cell(row,cond_col).value = "Good"
                            break
            self._write_log(wb, event.upper(), f"{event} on {asset_desc}: {note[:80]}")
            wb.save(MASTER_XL); wb.close()

    def write_ink(self, code, action, qty, big, sml, note=''):
        with self._lock:
            wb = openpyxl.load_workbook(MASTER_XL)
            wb[SHEET_INK].append([now(), code, INK_NAMES.get(code,code), action, qty, big, sml, note])
            self._write_log(wb, 'INK', f"Ink {action}: {qty}x {INK_NAMES.get(code,code)} ({code})")
            wb.save(MASTER_XL); wb.close()

    def get_transfer_log(self):
        """Return all transfer history entries."""
        with self._lock:
            wb  = openpyxl.load_workbook(MASTER_XL, data_only=True)
            ws  = wb[SHEET_HISTORY]
            out = []
            for row in list(ws.values)[1:]:
                if len(row) > 5 and str(row[5] if row[5] else '').strip() == 'transferred':
                    out.append({
                        'date':   str(row[0]) if row[0] else '',
                        'asset':  str(row[1]) if len(row)>1 and row[1] else '',
                        'serial': str(row[2]) if len(row)>2 and row[2] else '',
                        'from':   str(row[3]) if len(row)>3 and row[3] else '',
                        'to':     str(row[4]) if len(row)>4 and row[4] else '',
                        'note':   str(row[6]) if len(row)>6 and row[6] else '',
                        'by':     str(row[7]) if len(row)>7 and row[7] else '',
                    })
            wb.close()
            return list(reversed(out))


# ════════════════════════════════════════════
# JS API BRIDGE
# ════════════════════════════════════════════
class Api:
    def __init__(self, engine: ExcelEngine, cfg: ConfigEngine):
        self.engine = engine
        self.cfg    = cfg
        self.window = None

    def load_data(self):
        try:
            return json.dumps({"ok": True, "data": self.engine.load_all()})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def get_config(self):
        try:
            return json.dumps({"ok": True, "config": self.cfg.get()})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def add_config_item(self, key, value):
        try:
            ok = self.cfg.add_item(key, value)
            return json.dumps({"ok": ok, "msg": "Added" if ok else "Already exists"})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def remove_config_item(self, key, value):
        try:
            ok = self.cfg.remove_item(key, value)
            return json.dumps({"ok": ok})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def add_asset(self, data_json):
        try:
            data = json.loads(data_json)
            if not data.get('Equipment Type'):
                return json.dumps({"ok": False, "error": "Equipment Type is required"})
            if not data.get('Date Collected'):
                data['Date Collected'] = today()
            return json.dumps({"ok": True, "row": self.engine.add_asset(data)})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def update_asset(self, row_num, data_json):
        try:
            self.engine.update_asset(int(row_num), json.loads(data_json))
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def delete_asset(self, row_num, desc):
        try:
            self.engine.delete_asset(int(row_num), desc)
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def transfer_asset(self, row_num, asset_desc, serial, from_dept, to_dept, note, date, by='ICT Manager'):
        try:
            if not to_dept:
                return json.dumps({"ok": False, "error": "Target department is required"})
            if from_dept == to_dept:
                return json.dumps({"ok": False, "error": "Source and target departments must be different"})
            self.engine.transfer_asset(int(row_num), asset_desc, serial, from_dept, to_dept, note, date or today(), by)
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def get_transfer_log(self):
        try:
            return json.dumps({"ok": True, "data": self.engine.get_transfer_log()})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def record_replacement(self, rep_json):
        try:
            rep = json.loads(rep_json)
            if not rep.get('note'):
                return json.dumps({"ok": False, "error": "Replacement note is required"})
            self.engine.record_replacement(int(rep['oldRow']), rep['oldDesc'], rep)
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def write_history(self, asset_desc, serial, dept, event, note, date, cost='', by='ICT Manager'):
        try:
            self.engine.write_history_event(asset_desc, serial, dept, event, note, date, cost, by)
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def ink_action(self, code, action, qty, big, sml, note):
        try:
            if action == 'use':
                cur = self.engine.load_all()['ink'].get(code, {}).get('store', 0)
                if cur < int(qty):
                    return json.dumps({"ok": False, "error": f"Only {cur} in store"})
            self.engine.write_ink(code, action, int(qty), int(big), int(sml), note)
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def get_master_path(self):
        return MASTER_XL

    def open_excel(self):
        import subprocess
        try:
            if sys.platform == 'win32':    os.startfile(MASTER_XL)
            elif sys.platform == 'darwin': subprocess.Popen(['open', MASTER_XL])
            else:                          subprocess.Popen(['xdg-open', MASTER_XL])
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    # ── AUTH ──────────────────────────────────
    def is_pin_set(self):
        return json.dumps({"ok": True, "set": os.path.exists(AUTH_FILE)})

    def set_pin(self, pin):
        try:
            if len(pin.strip()) < 4:
                return json.dumps({"ok": False, "error": "PIN must be at least 4 characters"})
            salt = secrets.token_hex(16)
            h    = hashlib.sha256((salt + pin).encode()).hexdigest()
            with open(AUTH_FILE, 'w') as f:
                json.dump({"hash": h, "salt": salt}, f)
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def verify_pin(self, pin):
        try:
            if not os.path.exists(AUTH_FILE):
                return json.dumps({"ok": False, "error": "No PIN set"})
            with open(AUTH_FILE) as f:
                data = json.load(f)
            h = hashlib.sha256((data['salt'] + pin).encode()).hexdigest()
            return json.dumps({"ok": True, "valid": h == data['hash']})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    # ── PORTABLE APP BUILDER ──────────────────
    def build_installer(self):
        """Generate a self-contained setup script for portable deployment."""
        try:
            # Create setup script
            setup_bat = os.path.join(BASE_DIR, "INSTALL_ON_NEW_PC.bat")
            setup_sh  = os.path.join(BASE_DIR, "INSTALL_ON_NEW_PC.sh")
            
            bat_content = r"""@echo off
title St. Anne ICT -- Portable Setup
echo ================================================
echo  St. Anne Mission Hospital -- ICT Command Centre
echo  Portable Installer
echo ================================================
echo.
echo Step 1: Checking Python...
py -3.12 --version >nul 2>&1 && set PYTHON_CMD=py -3.12
if "%PYTHON_CMD%"=="" py -3.11 --version >nul 2>&1 && set PYTHON_CMD=py -3.11
if "%PYTHON_CMD%"=="" python --version >nul 2>&1 && set PYTHON_CMD=python
if "%PYTHON_CMD%"=="" (
    echo Python not found. Opening download page...
    start https://www.python.org/downloads/
    echo After install, re-run this script.
    pause
    exit /b 1
)
echo Python OK: %PYTHON_CMD%
echo.
echo Step 2: Installing required package...
%PYTHON_CMD% -m pip install openpyxl --quiet
echo.
echo Step 3: Launching application...
%PYTHON_CMD% main.py
pause
"""
            sh_content = """#!/bin/bash
echo "St. Anne Mission Hospital -- ICT Command Centre"
echo "Portable Installer"
echo ""
echo "Checking Python..."
if ! command -v python3 &> /dev/null; then
    echo "Python3 not found. Please install Python 3.10+"
    exit 1
fi
echo "Python OK"
echo ""
echo "Installing package..."
pip3 install openpyxl --quiet
echo ""
echo "Launching..."
python3 main.py
"""
            with open(setup_bat, 'w') as f: f.write(bat_content)
            with open(setup_sh, 'w') as f:  f.write(sh_content)
            try:
                import stat
                os.chmod(setup_sh, os.stat(setup_sh).st_mode | stat.S_IEXEC | stat.S_IXGRP | stat.S_IXOTH)
            except: pass
            
            # Create a README for portable use
            readme = os.path.join(BASE_DIR, "PORTABLE_INSTALL.txt")
            with open(readme, 'w') as f:
                f.write("""HOW TO INSTALL ON A NEW DEVICE
==============================

WINDOWS:
  1. Copy this entire folder to the new PC (USB stick works fine)
  2. Double-click INSTALL_ON_NEW_PC.bat
  3. It installs Python packages and launches automatically
  4. Next time: double-click start.bat

MAC / LINUX:
  1. Copy this folder to the new device
  2. Open Terminal, cd into this folder
  3. Run: bash INSTALL_ON_NEW_PC.sh
  4. Next time: python3 main.py

WHAT TO COPY:
  ✔ This entire folder including ICT_MASTER.xlsx
  ✔ All files: main.py, ui.html, auth.json, config.json
  ✔ Your data travels WITH the folder — no cloud, no server needed

SECURITY NOTE:
  - auth.json contains the access PIN (hashed, secure)
  - Do NOT share auth.json if you want to keep access restricted
  - To reset PIN: delete auth.json and restart the app

""")
            return json.dumps({"ok": True, "msg": f"Installer files created in: {BASE_DIR}"})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

# ════════════════════════════════════════════
# LAUNCH
# ════════════════════════════════════════════
BROWSER_BRIDGE = """
<script>
(function(){
  window.pywebview = window.pywebview || {};
  window.pywebview.api = new Proxy({}, {
    get: function(_, name) {
      return function() {
        return fetch('/api/' + encodeURIComponent(name), {
          method: 'POST',
          headers: {'Content-Type': 'application/json'},
          body: JSON.stringify({args: Array.prototype.slice.call(arguments)})
        }).then(function(r) { return r.text(); });
      };
    }
  });
  window.addEventListener('DOMContentLoaded', function(){
    window.dispatchEvent(new Event('pywebviewready'));
  });
})();
</script>
"""

def make_handler(api, ui_path):
    class AppHandler(http.server.BaseHTTPRequestHandler):
        def log_message(self, fmt, *args):
            return

        def _send(self, status, body, content_type="text/plain; charset=utf-8"):
            if isinstance(body, str):
                body = body.encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def do_GET(self):
            path = urllib.parse.urlparse(self.path).path
            if path in ("/", "/ui.html"):
                with open(ui_path, "r", encoding="utf-8") as f:
                    html = f.read()
                html = html.replace("</head>", BROWSER_BRIDGE + "\n</head>", 1)
                self._send(200, html, "text/html; charset=utf-8")
                return
            self._send(404, "Not found")

        def do_POST(self):
            path = urllib.parse.urlparse(self.path).path
            if not path.startswith("/api/"):
                self._send(404, "Not found")
                return
            name = urllib.parse.unquote(path.split("/api/", 1)[1])
            if name.startswith("_") or not hasattr(api, name):
                self._send(404, json.dumps({"ok": False, "error": "Unknown API method"}))
                return
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                result = getattr(api, name)(*payload.get("args", []))
                self._send(200, result)
            except Exception as e:
                self._send(500, json.dumps({"ok": False, "error": str(e)}))

    return AppHandler

class LocalServer(socketserver.ThreadingMixIn, http.server.HTTPServer):
    daemon_threads = True
    allow_reuse_address = True

def run_browser(api, ui_path):
    server = LocalServer(("127.0.0.1", 0), make_handler(api, ui_path))
    url = f"http://127.0.0.1:{server.server_port}/"
    print("St. Anne ICT Command Centre is running locally.")
    print("Open this address if the browser does not appear:")
    print(url)
    webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nClosing ICT Command Centre.")
    finally:
        server.server_close()

def run_desktop(api, ui_path):
    try:
        import webview
    except Exception as e:
        print("Desktop mode needs pywebview. Use start.bat for browser mode, or install pywebview first.")
        print(e)
        sys.exit(1)

    window = webview.create_window(
        title            = "St. Anne Mission Hospital — ICT Command Centre v2",
        url              = ui_path,
        js_api           = api,
        width            = 1440,
        height           = 880,
        min_size         = (1024, 680),
        background_color = "#0d1117",
        confirm_close    = True,
        text_select      = True,
    )
    api.window = window
    webview.start(debug=False, private_mode=False)

def main():
    engine = ExcelEngine()
    cfg    = ConfigEngine()
    api    = Api(engine, cfg)

    ui_path = os.path.join(BASE_DIR, 'ui.html')
    if not os.path.exists(ui_path):
        print("ERROR: ui.html not found in", BASE_DIR)
        sys.exit(1)

    if "--desktop" in sys.argv:
        run_desktop(api, ui_path)
    else:
        run_browser(api, ui_path)

if __name__ == "__main__":
    main()
