"""
St. Anne Mission Hospital — ICT Command Centre v2.0
Local browser app + HTML UI + Excel as live database
Enhancements: Custom departments/types, item transfers, repairs, replacements,
              classified role access, auto-audit logging, portable app builder
"""
import os as _os
_os.environ.setdefault("PYWEBVIEW_GUI", "edgechromium")
_os.environ.setdefault("WEBKIT_DISABLE_COMPOSITING_MODE", "1")
_os.environ.setdefault("PYWEBVIEW_NO_UIAUTOMATION", "1")

import json, os, sys, datetime, shutil, openpyxl, threading, re, hashlib, secrets
import http.server, socketserver, webbrowser, urllib.parse
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Portable EXE / Electron path resolution ───────────────────────────────────
# Three launch modes:
#   1. Electron:      env vars ICT_BUNDLE_DIR + ICT_DATA_DIR are set by main.js
#   2. PyInstaller:   sys._MEIPASS == bundle dir, sys.executable dir == data dir
#   3. Script (dev):  both dirs == directory of main.py

def _meipass():
    """Return the directory for read-only bundled assets (ui.html, seed Excel)."""
    if os.environ.get('ICT_BUNDLE_DIR'):
        return os.environ['ICT_BUNDLE_DIR']
    return getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))

def _datadir():
    """Return the directory for mutable data files (live Excel, auth, reports)."""
    if os.environ.get('ICT_DATA_DIR'):
        return os.environ['ICT_DATA_DIR']
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

BUNDLE_DIR = _meipass()   # read-only: ui.html, seed ICT_MASTER.xlsx
BASE_DIR   = _datadir()   # read-write: live data files

os.makedirs(BASE_DIR, exist_ok=True)

# On first run from a fresh install, seed the data files from the bundle
def _seed_file(name):
    dst = os.path.join(BASE_DIR, name)
    if not os.path.exists(dst):
        src = os.path.join(BUNDLE_DIR, name)
        if os.path.exists(src):
            shutil.copy2(src, dst)

for _f in ("ICT_MASTER.xlsx", "auth.json", "logo.png"):
    _seed_file(_f)

MASTER_XL  = os.path.join(BASE_DIR, "ICT_MASTER.xlsx")
BACKUP_DIR = os.path.join(BASE_DIR, "backups")
AUTH_FILE  = os.path.join(BASE_DIR, "auth.json")
CONFIG_FILE= os.path.join(BASE_DIR, "config.json")

SHEET_INV     = "Inventory"
SHEET_HISTORY = "History"
SHEET_REPLACE = "Replacements"
SHEET_INK     = "Ink"
SHEET_LOG     = "Audit Log"
SHEET_CONFIG  = "Config"
ALL_SHEETS    = [SHEET_INV, SHEET_HISTORY, SHEET_REPLACE, SHEET_INK, SHEET_LOG, SHEET_CONFIG]

INV_COLS = [
    "Date Collected","Purchase Date","Equipment Type","Brand/Model","Serial No.",
    "Assigned To","Department/Location","Condition/Status","OS/Firmware",
    "IP Address","MAC Address","CPU","RAM","Disk","Remarks","Transfer History"
]
INK_CODES    = ["BK","C","LC","M","LM","Y"]
INK_NAMES    = {"BK":"Black","C":"Cyan","LC":"Light Cyan","M":"Magenta","LM":"Light Magenta","Y":"Yellow"}
INK_DEFAULTS = {"BK":9,"C":3,"LC":5,"M":1,"LM":5,"Y":11}

# Default lists — users can extend these
DEFAULT_DEPTS = [
    "ICT Department","Administration","Outpatient","Inpatient","Pharmacy",
    "Laboratory","Radiology","Theatre","Maternity","Accounts","Records","Store"
]
DEFAULT_TYPES = [
    "Desktop Computer","Laptop","Printer","Scanner","Server","Switch","Router",
    "UPS","Monitor","Keyboard","Mouse","External HDD","Flash Drive","Projector",
    "Tablet","Phone","Photocopier","Camera","Other"
]
DEFAULT_CONDITIONS = [
    "New","Good","Fair","Fair (Needs Repair)","Needs Repair","Replaced","Disposed"
]

def today(): return datetime.date.today().strftime("%Y-%m-%d")
def now():   return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# ════════════════════════════════════════════
# CONFIG ENGINE
# ════════════════════════════════════════════
class ConfigEngine:
    def __init__(self):
        self._lock = threading.Lock()
        self._data = self._load()

    def _load(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE) as f:
                    return json.load(f)
            except: pass
        return {"departments": DEFAULT_DEPTS[:], "types": DEFAULT_TYPES[:], "conditions": DEFAULT_CONDITIONS[:]}

    def _save(self):
        with open(CONFIG_FILE, 'w') as f:
            json.dump(self._data, f, indent=2)

    def get(self):
        with self._lock:
            return dict(self._data)

    def add_item(self, key, value):
        with self._lock:
            v = value.strip()
            if v and v not in self._data.get(key, []):
                self._data.setdefault(key, []).append(v)
                self._save()
                return True
            return False

    def remove_item(self, key, value):
        with self._lock:
            lst = self._data.get(key, [])
            if value in lst:
                lst.remove(value)
                self._save()
                return True
            return False

# ════════════════════════════════════════════
# EXCEL ENGINE
# ════════════════════════════════════════════
class ExcelEngine:
    def __init__(self):
        self._lock = threading.Lock()
        self._ensure_workbook()

    def _init_headers(self, ws, name):
        hdrs = {
            SHEET_INV:     INV_COLS,
            SHEET_HISTORY: ["Date","Asset","Serial No.","From Department","To Department","Event","Description","By"],
            SHEET_REPLACE: ["Date Replaced","Old Asset","Old Serial","Department","Replaced With","New Serial","New Condition","Note","By"],
            SHEET_INK:     ["Timestamp","Colour Code","Colour Name","Action","Quantity","Big ICT","Small ICT","Note"],
            SHEET_LOG:     ["Timestamp","Action","Description","By"],
            SHEET_CONFIG:  ["Key","Value"],
        }
        ws.append(hdrs.get(name, []))
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", fgColor="1F2937")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

    def _set_widths(self, ws):
        widths = {
            "Date Collected":13,"Purchase Date":13,"Equipment Type":22,"Brand/Model":20,
            "Serial No.":22,"Assigned To":20,"Department/Location":24,"Condition/Status":22,
            "OS/Firmware":16,"IP Address":16,"MAC Address":18,"CPU":14,"RAM":12,"Disk":20,
            "Remarks":32,"Transfer History":40
        }
        if ws.max_row < 1: return
        h = [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        for i, col_name in enumerate(h, 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(col_name, 16)

    def _write_log(self, wb, action, desc, by="ICT Manager"):
        # Sanitise strings to avoid charmap codec errors on Windows
        def _safe(s): return str(s).encode("utf-8","replace").decode("utf-8") if s else ""
        wb[SHEET_LOG].append([now(), action, _safe(desc), _safe(by)])

    def _backup(self):
        os.makedirs(BACKUP_DIR, exist_ok=True)
        ts  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        dst = os.path.join(BACKUP_DIR, f"ICT_backup_{ts}.xlsx")
        shutil.copy2(MASTER_XL, dst)

    def _ensure_workbook(self):
        if not os.path.exists(MASTER_XL):
            self._create_fresh()
            return
        wb = openpyxl.load_workbook(MASTER_XL)
        OLD_DATA_SHEETS = ["work sheet","Work Sheet","worksheet","WorkSheet"]
        old_data_sheet  = next((s for s in OLD_DATA_SHEETS if s in wb.sheetnames), None)
        needs_migration = SHEET_INV not in wb.sheetnames and old_data_sheet is not None
        if needs_migration:
            os.makedirs(BACKUP_DIR, exist_ok=True)
            shutil.copy2(MASTER_XL, os.path.join(BACKUP_DIR, "ICT_original_before_migration.xlsx"))
            old_rows = self._read_old_rows(wb, old_data_sheet)
            for sname in list(wb.sheetnames):
                del wb[sname]
            for sheet in ALL_SHEETS:
                ws = wb.create_sheet(sheet)
                self._init_headers(ws, sheet)
            ws_inv = wb[SHEET_INV]
            for row in old_rows:
                # pad to new column count
                while len(row) < len(INV_COLS):
                    row.append('')
                ws_inv.append(row)
            self._set_widths(ws_inv)
            wb.save(MASTER_XL); wb.close()
            return
        changed = False
        for sheet in ALL_SHEETS:
            if sheet not in wb.sheetnames:
                ws = wb.create_sheet(sheet)
                self._init_headers(ws, sheet)
                changed = True
        # Add Transfer History column if missing
        if SHEET_INV in wb.sheetnames:
            ws = wb[SHEET_INV]
            if ws.max_row >= 1:
                h = [str(c.value).strip() if c.value else '' for c in ws[1]]
                if "Transfer History" not in h:
                    col = len(h) + 1
                    ws.cell(1, col).value = "Transfer History"
                    ws.cell(1, col).font  = Font(bold=True, color="FFFFFF", size=11)
                    ws.cell(1, col).fill  = PatternFill("solid", fgColor="1F2937")
                    changed = True
        if changed:
            wb.save(MASTER_XL)
        wb.close()

    def _read_old_rows(self, wb, sheet_name):
        ws   = wb[sheet_name]
        rows = list(ws.values)
        if not rows: return []
        old_h = [str(c).strip() if c else '' for c in rows[0]]
        def g(row, name, default=''):
            try:
                i = old_h.index(name)
                v = row[i] if i < len(row) else None
                if v is None: return default
                if isinstance(v, (datetime.date, datetime.datetime)):
                    return v.strftime('%Y-%m-%d')
                s = str(v).strip()
                if re.match(r'\d{4}-\d{2}-\d{2}[ T]', s): s = s[:10]
                return default if s.lower() in ['nan','none','nat','<na>'] else s
            except ValueError: return default
        def fix_ip(ip): return (ip or '').replace('192.068.','192.168.').replace('192.0.168.','192.168.')
        result = []
        for row in rows[1:]:
            if not any(v for v in row): continue
            ram  = ' '.join(p for p in [g(row,'RAM Size'),g(row,'RAM Type')] if p)
            disk = ' '.join(p for p in [g(row,'Disk Size'),g(row,'Disk Type'),g(row,'Disk Health')] if p)
            result.append([
                g(row,'Date Collected'), g(row,'Purchase Date'), g(row,'Equipment Type'),
                g(row,'Brand/Model'), g(row,'Serial No.'), g(row,'Assigned To'),
                g(row,'Department/Location').strip(), g(row,'Condition/Status') or 'Good',
                g(row,'OS/Firmware'), fix_ip(g(row,'IP Address')), g(row,'MAC Address'),
                g(row,'CPU'), ram, disk, g(row,'Remarks'), ''
            ])
        return result

    def _create_fresh(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for s in ALL_SHEETS:
            self._init_headers(wb.create_sheet(s), s)
        wb.save(MASTER_XL)

    def _clean(self, v):
        if v is None: return ''
        if isinstance(v, (datetime.date, datetime.datetime)): return v.strftime('%Y-%m-%d')
        s = str(v).strip()
        if re.match(r'\d{4}-\d{2}-\d{2}[ T]', s): s = s[:10]
        return '' if s.lower() in ['nan','none','nat','<na>'] else s

    # ── READ ─────────────────────────────────
    def load_all(self):
        with self._lock:
            wb   = openpyxl.load_workbook(MASTER_XL, data_only=True)
            inv  = self._read_inv(wb)
            ink  = self._read_ink(wb)
            log  = self._read_log(wb)
            hist = self._read_hist(wb)
            wb.close()
            return {"inventory": inv, "ink": ink, "log": log, "history": hist}

    def log_session(self, action, desc, by="ICT Manager"):
        """Write a standalone audit log entry without touching other sheets."""
        with self._lock:
            wb = openpyxl.load_workbook(MASTER_XL)
            self._write_log(wb, action, desc, by)
            wb.save(MASTER_XL); wb.close()

    def _read_inv(self, wb):
        ws   = wb[SHEET_INV]
        rows = list(ws.values)
        if len(rows) < 2: return []
        h   = [str(c).strip() if c else '' for c in rows[0]]
        out = []
        for i, row in enumerate(rows[1:], start=2):
            d = {'_row': i}
            for j, col in enumerate(h):
                d[col] = self._clean(row[j] if j < len(row) else None)
            out.append(d)
        return out

    def _read_ink(self, wb):
        ws    = wb[SHEET_INK]
        rows  = list(ws.values)
        # state: store = storeroom qty, ictRoom = ICT room qty (deployed from store)
        state = {k: {"store": INK_DEFAULTS[k], "ictRoom": 0} for k in INK_CODES}
        for row in rows[1:]:
            if not row or not row[1]: continue
            code   = str(row[1]).strip()
            action = str(row[3]).strip() if len(row)>3 and row[3] else ''
            try: qty = int(float(str(row[4]))) if len(row)>4 and row[4] else 0
            except: qty = 0
            # bigIct column now stores ictRoom qty snapshot (written on move_to_ict/use)
            try: ict_snap = int(float(str(row[5]))) if len(row)>5 and row[5] not in (None,'','None') else None
            except: ict_snap = None
            if code in state:
                if action == 'restock':
                    state[code]['store'] += qty
                elif action == 'move_to_ict':
                    # Move from store to ICT room
                    state[code]['store']   = max(0, state[code]['store'] - qty)
                    state[code]['ictRoom'] += qty
                elif action == 'use':
                    # Use from ICT room first, then store
                    from_ict = min(qty, state[code]['ictRoom'])
                    remaining = qty - from_ict
                    state[code]['ictRoom'] = max(0, state[code]['ictRoom'] - from_ict)
                    state[code]['store']   = max(0, state[code]['store'] - remaining)
                elif action == 'set_printer':
                    pass
                # If a snapshot of ictRoom was recorded, trust it (keeps state consistent)
                if ict_snap is not None and action in ('move_to_ict','use','set_printer'):
                    state[code]['ictRoom'] = ict_snap
        return state

    def _read_log(self, wb):
        ws  = wb[SHEET_LOG]
        out = []
        for row in list(ws.values)[1:]:
            out.append({'time':str(row[0]) if row[0] else '','action':str(row[1]) if len(row)>1 and row[1] else '','desc':str(row[2]) if len(row)>2 and row[2] else ''})
        return list(reversed(out))

    def _read_hist(self, wb):
        ws  = wb[SHEET_HISTORY]
        out = []
        for row in list(ws.values)[1:]:
            out.append({
                'date':  str(row[0]) if row[0] else '',
                'asset': str(row[1]) if len(row)>1 and row[1] else '',
                'serial':str(row[2]) if len(row)>2 and row[2] else '',
                'from':  str(row[3]) if len(row)>3 and row[3] else '',
                'to':    str(row[4]) if len(row)>4 and row[4] else '',
                'event': str(row[5]) if len(row)>5 and row[5] else '',
                'note':  str(row[6]) if len(row)>6 and row[6] else '',
                'by':    str(row[7]) if len(row)>7 and row[7] else '',
            })
        return list(reversed(out))

    # ── WRITE ────────────────────────────────
    def add_asset(self, data):
        with self._lock:
            self._backup()
            wb = openpyxl.load_workbook(MASTER_XL)
            ws = wb[SHEET_INV]
            row_data = [data.get(c, '') for c in INV_COLS]
            ws.append(row_data)
            new_row = ws.max_row
            self._set_widths(ws)
            self._write_log(wb, 'ADD', f"Added: {data.get('Equipment Type','')} — {data.get('Brand/Model','')} | {data.get('Department/Location','')}")
            wb.save(MASTER_XL); wb.close()
            return new_row

    def update_asset(self, row_num, data):
        with self._lock:
            self._backup()
            wb = openpyxl.load_workbook(MASTER_XL)
            ws = wb[SHEET_INV]
            h  = [ws.cell(1,c).value for c in range(1,len(INV_COLS)+2)]
            for j, col in enumerate(h, 1):
                if col in data:
                    ws.cell(row=row_num, column=j).value = data[col]
            self._write_log(wb, 'EDIT', f"Edited row {row_num}: {data.get('Equipment Type','')} — {data.get('Brand/Model','')}")
            wb.save(MASTER_XL); wb.close()

    def delete_asset(self, row_num, desc):
        with self._lock:
            self._backup()
            wb = openpyxl.load_workbook(MASTER_XL)
            wb[SHEET_INV].delete_rows(row_num)
            self._write_log(wb, 'DELETE', f"Deleted: {desc}")
            wb.save(MASTER_XL); wb.close()

    def transfer_asset(self, row_num, asset_desc, serial, from_dept, to_dept, note, date, by='ICT Manager'):
        """Transfer an asset to another department — updates Inventory + logs history."""
        with self._lock:
            self._backup()
            wb     = openpyxl.load_workbook(MASTER_XL)
            ws_inv = wb[SHEET_INV]
            h      = [ws_inv.cell(1,c).value for c in range(1, ws_inv.max_column+1)]
            def col_idx(name):
                try: return h.index(name)+1
                except: return None
            dept_col = col_idx("Department/Location")
            xfer_col = col_idx("Transfer History")
            if dept_col:
                ws_inv.cell(row=row_num, column=dept_col).value = to_dept
            if xfer_col:
                old_xfer = ws_inv.cell(row=row_num, column=xfer_col).value or ''
                entry = f"[{date}] {from_dept} → {to_dept}"
                if note: entry += f": {note}"
                ws_inv.cell(row=row_num, column=xfer_col).value = (str(old_xfer) + " | " + entry).strip(" |")
            # History
            by_val = by if by and by.strip() else "ICT Manager"
            wb[SHEET_HISTORY].append([date, asset_desc, serial, from_dept, to_dept, "transferred",
                f"Transferred from {from_dept} to {to_dept}. Note: {note}", by_val])
            self._write_log(wb, 'TRANSFER', f"Transferred {asset_desc} (S/N:{serial}) from {from_dept} → {to_dept}")
            wb.save(MASTER_XL); wb.close()

    def record_replacement(self, old_row, old_desc, rep):
        with self._lock:
            self._backup()
            wb     = openpyxl.load_workbook(MASTER_XL)
            ws_inv = wb[SHEET_INV]
            h      = [ws_inv.cell(1,c).value for c in range(1, len(INV_COLS)+2)]
            def col_idx(name):
                try: return h.index(name)+1
                except: return None
            ci = col_idx("Condition/Status"); ri = col_idx("Remarks")
            if ci: ws_inv.cell(row=old_row, column=ci).value = "Replaced"
            if ri:
                old_rem = ws_inv.cell(row=old_row, column=ri).value or ''
                ws_inv.cell(row=old_row, column=ri).value = (str(old_rem)+f" | REPLACED {rep['date']}: {rep['note']}").strip(" |")
            ws_inv.append([{
                "Date Collected":rep['date'],"Purchase Date":rep['date'],
                "Equipment Type":rep.get('newType',''),"Brand/Model":rep.get('newBrand',''),
                "Serial No.":rep.get('newSerial',''),"Assigned To":rep.get('assigned',''),
                "Department/Location":rep.get('dept',''),"Condition/Status":rep.get('newCond','New'),
                "OS/Firmware":"","IP Address":rep.get('ip',''),"MAC Address":"","CPU":"","RAM":"","Disk":"",
                "Remarks":f"Replaced: {old_desc} on {rep['date']}. {rep['note']}","Transfer History":""
            }.get(c,'') for c in INV_COLS])
            self._set_widths(ws_inv)
            wb[SHEET_REPLACE].append([rep['date'],old_desc,rep.get('oldSerial',''),rep.get('dept',''),
                f"{rep.get('newType','')} {rep.get('newBrand','')}".strip(),rep.get('newSerial',''),
                rep.get('newCond','New'),rep['note'],"ICT Manager"])
            wb[SHEET_HISTORY].append([rep['date'],old_desc,rep.get('oldSerial',''),rep.get('dept',''),
                rep.get('dept',''),"replaced",
                f"Replaced with: {rep.get('newType','')} {rep.get('newBrand','')} S/N:{rep.get('newSerial','')}. Note: {rep['note']}","ICT Manager"])
            self._write_log(wb,'REPLACE',f"Replaced {old_desc} → {rep.get('newBrand','')} on {rep['date']}")
            wb.save(MASTER_XL); wb.close()

    def write_history_event(self, asset_desc, serial, dept, event, note, date, cost='', by='ICT Manager'):
        with self._lock:
            wb = openpyxl.load_workbook(MASTER_XL)
            by_val = by if by and by.strip() else "ICT Manager"
            wb[SHEET_HISTORY].append([date, asset_desc, serial, dept, dept, event, note, by_val])

            # For events that change asset state, update the Inventory row too
            STATUS_CHANGE_EVENTS = {'repaired', 'disposed', 'replaced', 'upgraded'}
            if event in STATUS_CHANGE_EVENTS or event == 'reassigned':
                ws = wb[SHEET_INV]
                h  = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
                def col_idx(name):
                    try: return h.index(name)+1
                    except ValueError: return None
                ser_col   = col_idx("Serial No.")
                cond_col  = col_idx("Condition/Status")
                assign_col= col_idx("Assigned To")
                dept_col  = col_idx("Department/Location")

                if ser_col:
                    for row in range(2, ws.max_row+1):
                        if str(ws.cell(row, ser_col).value or '').strip() == serial.strip():
                            cur_cond = str(ws.cell(row, cond_col).value or '') if cond_col else ''
                            if event == 'repaired' and cond_col:
                                if 'Repair' in cur_cond or cur_cond in ('Fair (Needs Repair)', 'Needs Repair'):
                                    ws.cell(row, cond_col).value = 'Good'
                            elif event == 'disposed' and cond_col:
                                ws.cell(row, cond_col).value = 'Disposed'
                            elif event == 'replaced' and cond_col:
                                ws.cell(row, cond_col).value = 'Replaced'
                            elif event == 'upgraded' and cond_col:
                                # Mark as Good after upgrade
                                ws.cell(row, cond_col).value = 'Good'
                            elif event == 'reassigned':
                                # Extract new assignee from note if present e.g. "Reassigned to John"
                                import re as _re
                                m = _re.search(r'(?:to|→)\s*(.+?)(?:\s*\[|$)', note, _re.IGNORECASE)
                                if m and assign_col:
                                    ws.cell(row, assign_col).value = m.group(1).strip()
                            break

            self._write_log(wb, event.upper(), f"{event} on {asset_desc}: {note[:80]}")
            wb.save(MASTER_XL); wb.close()

    def write_ink(self, code, action, qty, ict_room_after, note=''):
        with self._lock:
            wb = openpyxl.load_workbook(MASTER_XL)
            # Columns: Timestamp, Code, Name, Action, Qty, ICT Room (snapshot), _, Note
            wb[SHEET_INK].append([now(), code, INK_NAMES.get(code,code), action, qty, ict_room_after, 0, note])
            if action == 'restock':
                self._write_log(wb, 'INK', f"Ink restocked to store: {qty}x {INK_NAMES.get(code,code)} ({code})")
            elif action == 'move_to_ict':
                self._write_log(wb, 'INK', f"Ink moved Store→ICT Room: {qty}x {INK_NAMES.get(code,code)} ({code})")
            elif action == 'use':
                self._write_log(wb, 'INK', f"Ink used: {qty}x {INK_NAMES.get(code,code)} ({code})")
            else:
                self._write_log(wb, 'INK', f"Ink {action}: {qty}x {INK_NAMES.get(code,code)} ({code})")
            wb.save(MASTER_XL); wb.close()

    def get_transfer_log(self):
        """Return all transfer history entries."""
        with self._lock:
            wb  = openpyxl.load_workbook(MASTER_XL, data_only=True)
            ws  = wb[SHEET_HISTORY]
            out = []
            for row in list(ws.values)[1:]:
                if len(row) > 5 and str(row[5] if row[5] else '').strip() == 'transferred':
                    out.append({
                        'date':   str(row[0]) if row[0] else '',
                        'asset':  str(row[1]) if len(row)>1 and row[1] else '',
                        'serial': str(row[2]) if len(row)>2 and row[2] else '',
                        'from':   str(row[3]) if len(row)>3 and row[3] else '',
                        'to':     str(row[4]) if len(row)>4 and row[4] else '',
                        'note':   str(row[6]) if len(row)>6 and row[6] else '',
                        'by':     str(row[7]) if len(row)>7 and row[7] else '',
                    })
            wb.close()
            return list(reversed(out))


# ════════════════════════════════════════════
# JS API BRIDGE
# ════════════════════════════════════════════
class Api:
    def __init__(self, engine: ExcelEngine, cfg: ConfigEngine):
        self.engine = engine
        self.cfg    = cfg
        self.window = None

    def load_data(self):
        try:
            return json.dumps({"ok": True, "data": self.engine.load_all()})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def get_config(self):
        try:
            return json.dumps({"ok": True, "config": self.cfg.get()})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def add_config_item(self, key, value):
        try:
            ok = self.cfg.add_item(key, value)
            return json.dumps({"ok": ok, "msg": "Added" if ok else "Already exists"})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def remove_config_item(self, key, value):
        try:
            ok = self.cfg.remove_item(key, value)
            return json.dumps({"ok": ok})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def add_asset(self, data_json):
        try:
            data = json.loads(data_json)
            if not data.get('Equipment Type'):
                return json.dumps({"ok": False, "error": "Equipment Type is required"})
            if not data.get('Date Collected'):
                data['Date Collected'] = today()
            return json.dumps({"ok": True, "row": self.engine.add_asset(data)})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def update_asset(self, row_num, data_json):
        try:
            self.engine.update_asset(int(row_num), json.loads(data_json))
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def delete_asset(self, row_num, desc):
        try:
            self.engine.delete_asset(int(row_num), desc)
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def transfer_asset(self, row_num, asset_desc, serial, from_dept, to_dept, note, date, by='ICT Manager'):
        try:
            if not to_dept:
                return json.dumps({"ok": False, "error": "Target department is required"})
            if from_dept == to_dept:
                return json.dumps({"ok": False, "error": "Source and target departments must be different"})
            self.engine.transfer_asset(int(row_num), asset_desc, serial, from_dept, to_dept, note, date or today(), by)
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def get_transfer_log(self):
        try:
            return json.dumps({"ok": True, "data": self.engine.get_transfer_log()})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def record_replacement(self, rep_json):
        try:
            rep = json.loads(rep_json)
            if not rep.get('note'):
                return json.dumps({"ok": False, "error": "Replacement note is required"})
            self.engine.record_replacement(int(rep['oldRow']), rep['oldDesc'], rep)
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def write_history(self, asset_desc, serial, dept, event, note, date, cost='', by='ICT Manager'):
        try:
            self.engine.write_history_event(asset_desc, serial, dept, event, note, date, cost, by)
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def ink_action(self, code, action, qty, ict_room_after=0, note=''):
        try:
            qty = int(qty) if qty else 0
            ict_room_after = int(ict_room_after) if ict_room_after else 0
            if action == 'use':
                state = self.engine.load_all()['ink'].get(code, {})
                total_avail = state.get('ictRoom', 0) + state.get('store', 0)
                if total_avail < qty:
                    return json.dumps({"ok": False, "error": f"Only {total_avail} available (ICT Room + Store)"})
            elif action == 'move_to_ict':
                state = self.engine.load_all()['ink'].get(code, {})
                if state.get('store', 0) < qty:
                    return json.dumps({"ok": False, "error": f"Only {state.get('store',0)} in store"})
            self.engine.write_ink(code, action, qty, ict_room_after, note)
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def log_action(self, action, desc, by="ICT Manager"):
        """Client-side initiated audit log entry."""
        try:
            self.engine.log_session(action, desc, by)
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def generate_pdf(self, dept_filter='', type_filter='', date_from='', date_to=''):
        """Generate PDF report, return base64 content for browser download."""
        try:
            import os, base64
            out_dir = os.path.join(BASE_DIR, 'reports')
            os.makedirs(out_dir, exist_ok=True)
            ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            fname = f'ICT_Report_{ts}.pdf'
            out_path = os.path.join(out_dir, fname)
            logo_path = ''
            for lname in ['logo.png','logo.jpg','logo.jpeg','LOGO.PNG','LOGO.JPG']:
                candidate = os.path.join(BASE_DIR, lname)
                if os.path.exists(candidate):
                    logo_path = candidate
                    break
            result = generate_pdf_report(self.engine, out_path, dept_filter, type_filter, logo_path, date_from, date_to)
            if result['ok']:
                result['filename'] = fname
                with open(out_path, 'rb') as f:
                    result['b64'] = base64.b64encode(f.read()).decode('ascii')
            return json.dumps(result)
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def get_logo_path(self):
        """Return a URL to the logo — HTTP in Electron/browser mode, file:/// in desktop mode."""
        for lname in ['logo.png','logo.jpg','logo.jpeg','LOGO.PNG','LOGO.JPG']:
            p = os.path.join(BASE_DIR, lname)
            if os.path.exists(p):
                # In Electron/browser mode the HTTP server is running — serve via /logo
                port = os.environ.get('ICT_HTTP_PORT')
                if port:
                    return f'http://127.0.0.1:{port}/logo'
                # Desktop (pywebview) mode — file:// works fine
                uri = p.replace(os.sep, '/'); return 'file:///' + uri
        return ''

    def get_master_path(self):
        return MASTER_XL

    def open_excel(self):
        import subprocess
        try:
            if sys.platform == 'win32':    os.startfile(MASTER_XL)
            elif sys.platform == 'darwin': subprocess.Popen(['open', MASTER_XL])
            else:                          subprocess.Popen(['xdg-open', MASTER_XL])
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    # ── AUTH ──────────────────────────────────
    def is_pin_set(self):
        return json.dumps({"ok": True, "set": os.path.exists(AUTH_FILE)})

    def set_pin(self, pin):
        try:
            if len(pin.strip()) < 4:
                return json.dumps({"ok": False, "error": "PIN must be at least 4 characters"})
            salt = secrets.token_hex(16)
            h    = hashlib.sha256((salt + pin).encode()).hexdigest()
            with open(AUTH_FILE, 'w') as f:
                json.dump({"hash": h, "salt": salt}, f)
            return json.dumps({"ok": True})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def verify_pin(self, pin):
        try:
            if not os.path.exists(AUTH_FILE):
                return json.dumps({"ok": False, "error": "No PIN set"})
            with open(AUTH_FILE) as f:
                data = json.load(f)
            h = hashlib.sha256((data['salt'] + pin).encode()).hexdigest()
            return json.dumps({"ok": True, "valid": h == data['hash']})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    # ── PORTABLE APP BUILDER ──────────────────
    def build_installer(self):
        """Generate robust self-contained installer scripts for portable deployment."""
        try:
            setup_bat = os.path.join(BASE_DIR, "INSTALL_ON_NEW_PC.bat")
            setup_sh  = os.path.join(BASE_DIR, "INSTALL_ON_NEW_PC.sh")
            start_bat = os.path.join(BASE_DIR, "start.bat")

            bat_content = (
                "@echo off\r\n"
                "title St. Anne Mission Hospital - ICT Command Centre Setup\r\n"
                "color 0A\r\n"
                "echo.\r\n"
                "echo  =====================================================\r\n"
                "echo   St. Anne Mission Hospital - ICT Command Centre\r\n"
                "echo   Portable Installer  ^|  Serve With Love\r\n"
                "echo  =====================================================\r\n"
                "echo.\r\n"
                "\r\n"
                ":: ── Step 1: Find Python ──\r\n"
                "set PYTHON_CMD=\r\n"
                "for %%C in (python py python3) do (\r\n"
                "  if defined PYTHON_CMD goto :found\r\n"
                "  %%C --version >nul 2>&1 && set PYTHON_CMD=%%C\r\n"
                ")\r\n"
                ":found\r\n"
                "if not defined PYTHON_CMD (\r\n"
                "  echo  [!] Python not found.\r\n"
                "  echo  [>] Opening Python download page...\r\n"
                "  start https://www.python.org/downloads/\r\n"
                "  echo  [!] Install Python 3.10 or higher, then run this script again.\r\n"
                "  echo  [!] TIP: Tick 'Add Python to PATH' during installation!\r\n"
                "  pause\r\n"
                "  exit /b 1\r\n"
                ")\r\n"
                "echo  [+] Python found: %PYTHON_CMD%\r\n"
                "\r\n"
                ":: ── Step 2: Upgrade pip silently ──\r\n"
                "echo  [~] Updating pip...\r\n"
                "%PYTHON_CMD% -m pip install --upgrade pip --quiet --no-warn-script-location 2>nul\r\n"
                "\r\n"
                ":: ── Step 3: Install required packages ──\r\n"
                "echo  [~] Installing required packages...\r\n"
                "%PYTHON_CMD% -m pip install openpyxl reportlab --quiet --no-warn-script-location\r\n"
                "if errorlevel 1 (\r\n"
                "  echo  [!] Package installation failed. Check your internet connection.\r\n"
                "  pause\r\n"
                "  exit /b 1\r\n"
                ")\r\n"
                "echo  [+] Packages installed successfully.\r\n"
                "\r\n"
                ":: ── Step 4: Launch app ──\r\n"
                "echo  [>] Launching ICT Command Centre...\r\n"
                "echo.\r\n"
                "%PYTHON_CMD% main.py\r\n"
                "pause\r\n"
            )

            start_bat_content = (
                "@echo off\r\n"
                "title St. Anne Mission Hospital - ICT Command Centre\r\n"
                "cd /d \"%~dp0\"\r\n"
                "for %%C in (python py python3) do (\r\n"
                "  %%C --version >nul 2>&1 && set PYTHON_CMD=%%C && goto :run\r\n"
                ")\r\n"
                "echo Python not found. Please run INSTALL_ON_NEW_PC.bat first.\r\n"
                "pause\r\n"
                "exit /b 1\r\n"
                ":run\r\n"
                "start \"\" /B %PYTHON_CMD% main.py\r\n"
                "exit\r\n"
            )

            sh_content = (
                "#!/usr/bin/env bash\n"
                "set -e\n"
                "echo\n"
                "echo ' ====================================================='\n"
                "echo '  St. Anne Mission Hospital — ICT Command Centre'\n"
                "echo '  Portable Installer  |  Serve With Love'\n"
                "echo ' ====================================================='\n"
                "echo\n"
                "# Find Python\n"
                "PYTHON_CMD=''\n"
                "for cmd in python3 python python3.12 python3.11 python3.10; do\n"
                "  if command -v $cmd &>/dev/null; then PYTHON_CMD=$cmd; break; fi\n"
                "done\n"
                "if [ -z \"$PYTHON_CMD\" ]; then\n"
                "  echo '[!] Python 3 not found. Install Python 3.10+ and retry.'\n"
                "  exit 1\n"
                "fi\n"
                "echo \"[+] Python found: $PYTHON_CMD\"\n"
                "echo '[~] Installing packages...'\n"
                "$PYTHON_CMD -m pip install openpyxl reportlab --quiet\n"
                "echo '[+] Packages ready.'\n"
                "echo '[>] Launching...'\n"
                "$PYTHON_CMD main.py\n"
            )

            readme_content = (
                "ST. ANNE MISSION HOSPITAL — ICT COMMAND CENTRE\n"
                "HOW TO INSTALL ON A NEW DEVICE\n"
                "===============================================\n\n"
                "WINDOWS:\n"
                "  1. Copy this entire folder to the new PC (USB stick is fine)\n"
                "  2. Double-click  INSTALL_ON_NEW_PC.bat\n"
                "     • It will auto-install Python packages and launch the app\n"
                "  3. Next time, just double-click  start.bat\n\n"
                "MAC / LINUX:\n"
                "  1. Copy this folder to the new device\n"
                "  2. Open Terminal and cd into this folder\n"
                "  3. Run:  bash INSTALL_ON_NEW_PC.sh\n"
                "  4. Next time:  python3 main.py\n\n"
                "REQUIREMENTS:\n"
                "  • Python 3.10 or higher  (https://www.python.org/downloads/)\n"
                "  • Internet connection for first-time package install\n"
                "  • Packages: openpyxl, reportlab (installed automatically)\n\n"
                "WHAT TO COPY TO NEW PC:\n"
                "  main.py          — the application\n"
                "  ui.html          — the interface\n"
                "  ICT_MASTER.xlsx  — your live database (all your data)\n"
                "  auth.json        — your access PIN (hashed)\n"
                "  config.json      — your custom departments/types (if exists)\n"
                "  logo.png         — your hospital logo (if you have one)\n"
                "  start.bat        — quick-start launcher for Windows\n\n"
                "LOGO:\n"
                "  Place a file named logo.png in this folder.\n"
                "  It will appear on every PDF report page header.\n\n"
                "SECURITY:\n"
                "  • auth.json holds your PIN hash — do not share if access-restricted\n"
                "  • To reset PIN: delete auth.json and restart the app\n\n"
                "SUPPORT:\n"
                "  ICT Department — St. Anne Mission Hospital\n"
                "  Serve With Love\n"
            )

            with open(setup_bat, 'w', newline='', encoding='utf-8') as f: f.write(bat_content)
            with open(start_bat, 'w', newline='', encoding='utf-8') as f: f.write(start_bat_content)
            with open(setup_sh,  'w', newline='\n', encoding='utf-8') as f: f.write(sh_content)
            with open(os.path.join(BASE_DIR, 'PORTABLE_README.txt'), 'w', encoding='utf-8') as f: f.write(readme_content)

            try:
                import stat as _stat
                os.chmod(setup_sh, os.stat(setup_sh).st_mode | _stat.S_IEXEC | _stat.S_IXGRP | _stat.S_IXOTH)
            except: pass

            self.engine.log_session('INSTALLER', 'Portable installer scripts generated')
            return json.dumps({"ok": True, "msg": f"Installer files created in:\n{BASE_DIR}\n\nFiles: INSTALL_ON_NEW_PC.bat / .sh  +  start.bat  +  PORTABLE_README.txt"})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

# ════════════════════════════════════════════
# LAUNCH
# ════════════════════════════════════════════
BROWSER_BRIDGE = """
<script>
(function(){
  window.pywebview = window.pywebview || {};
  window.pywebview.api = new Proxy({}, {
    get: function(_, name) {
      return function() {
        return fetch('/api/' + encodeURIComponent(name), {
          method: 'POST',
          headers: {'Content-Type': 'application/json'},
          body: JSON.stringify({args: Array.prototype.slice.call(arguments)})
        }).then(function(r) { return r.text(); });
      };
    }
  });
  // Mark that we are running in browser mode (not pywebview desktop)
  window.__IS_BROWSER_MODE__ = true;
  // Wait for DOMContentLoaded, then yield with setTimeout(0) so that
  // all synchronous <script> blocks at the bottom of <body> have had a
  // chance to register their 'pywebviewready' listeners before we fire.
  function dispatchReady() {
    window.dispatchEvent(new Event('pywebviewready'));
  }
  if (document.readyState === 'loading') {
    window.addEventListener('DOMContentLoaded', function(){ setTimeout(dispatchReady, 0); });
  } else {
    setTimeout(dispatchReady, 0);
  }
})();
</script>
"""

def make_handler(api, ui_path, browser_mode=False):
    class AppHandler(http.server.BaseHTTPRequestHandler):
        def log_message(self, fmt, *args):
            return

        def _send(self, status, body, content_type="text/plain; charset=utf-8"):
            if isinstance(body, str):
                body = body.encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def do_GET(self):
            path = urllib.parse.urlparse(self.path).path
            if path in ("/", "/ui.html"):
                with open(ui_path, "r", encoding="utf-8") as f:
                    html = f.read()
                # Only inject the fetch-proxy bridge in browser mode.
                # In desktop (pywebview) mode pywebview injects its own native
                # js_api bridge — injecting ours would overwrite it.
                if browser_mode:
                    html = html.replace("</head>", BROWSER_BRIDGE + "\n</head>", 1)
                self._send(200, html, "text/html; charset=utf-8")
                return
            # Serve logo image over HTTP (needed in Electron mode — file:// blocked)
            if path == "/logo":
                for lname in ['logo.png','logo.jpg','logo.jpeg','LOGO.PNG','LOGO.JPG']:
                    lp = os.path.join(BASE_DIR, lname)
                    if os.path.isfile(lp):
                        ext = os.path.splitext(lname)[1].lower()
                        mime = 'image/jpeg' if ext in ('.jpg','.jpeg') else 'image/png'
                        with open(lp, 'rb') as lf:
                            data = lf.read()
                        self.send_response(200)
                        self.send_header('Content-Type', mime)
                        self.send_header('Content-Length', str(len(data)))
                        self.end_headers()
                        self.wfile.write(data)
                        return
                self._send(404, "No logo")
                return
            # Serve generated PDF reports for browser download
            if path.startswith("/reports/"):
                fname = urllib.parse.unquote(path[len("/reports/"):])
                # Sanitise: no path traversal
                if fname and "/" not in fname and "\\" not in fname:
                    fpath = os.path.join(BASE_DIR, "reports", fname)
                    if os.path.isfile(fpath):
                        with open(fpath, "rb") as pf:
                            data = pf.read()
                        self.send_response(200)
                        self.send_header("Content-Type", "application/pdf")
                        self.send_header("Content-Length", str(len(data)))
                        self.send_header("Content-Disposition",
                                         f'inline; filename="{fname}"')
                        self.end_headers()
                        self.wfile.write(data)
                        return
            self._send(404, "Not found")

        def do_POST(self):
            path = urllib.parse.urlparse(self.path).path
            if not path.startswith("/api/"):
                self._send(404, "Not found")
                return
            name = urllib.parse.unquote(path.split("/api/", 1)[1])
            if name.startswith("_") or not hasattr(api, name):
                self._send(404, json.dumps({"ok": False, "error": "Unknown API method"}))
                return
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                result = getattr(api, name)(*payload.get("args", []))
                self._send(200, result)
            except Exception as e:
                self._send(500, json.dumps({"ok": False, "error": str(e)}))

    return AppHandler

class LocalServer(socketserver.ThreadingMixIn, http.server.HTTPServer):
    daemon_threads = True
    allow_reuse_address = True

def run_browser(api, ui_path, electron_mode=False):
    # electron_mode: bridge injected by preload.js; serve raw html, no webbrowser.open
    server = LocalServer(("127.0.0.1", 0), make_handler(api, ui_path, browser_mode=not electron_mode))
    port   = server.server_port
    url    = f"http://127.0.0.1:{port}/"
    # Expose port so get_logo_path() can build an HTTP URL for the logo
    os.environ['ICT_HTTP_PORT'] = str(port)
    if electron_mode:
        print(f"ELECTRON_PORT={port}", flush=True)
    else:
        print("St. Anne ICT Command Centre is running locally.")
        print("Open this address if the browser does not appear:")
        print(url)
        webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nClosing ICT Command Centre.")
    finally:
        server.server_close()

def run_desktop(api, ui_path):
    try:
        import webview
    except Exception as e:
        print("Desktop mode needs pywebview. Use start.bat for browser mode, or install pywebview first.")
        print(e)
        sys.exit(1)

    # Desktop mode: pywebview loads ui.html directly via file:// and exposes
    # js_api as window.pywebview.api natively. No HTTP server or bridge injection needed.
    window = webview.create_window(
        title            = "St. Anne Mission Hospital — ICT Command Centre v2",
        url              = ui_path,
        js_api           = api,
        width            = 1440,
        height           = 880,
        min_size         = (1024, 680),
        background_color = "#0d1117",
        confirm_close    = True,
        text_select      = True,
    )
    api.window = window
    webview.start(debug=False, private_mode=False)

# ════════════════════════════════════════════
# PROFESSIONAL PDF REPORT ENGINE  (2-page max)
# ════════════════════════════════════════════
def generate_pdf_report(engine: "ExcelEngine", output_path: str,
                        dept_filter: str = "", type_filter: str = "",
                        logo_path: str = "", date_from: str = "", date_to: str = "") -> dict:
    """
    Generate a concise 2-page professional ICT Asset Report.
    Page 1 : Header + KPI strip + Executive Summary + Equipment table + Department table
    Page 2 : (only if needed) Attention list + Recent Activity + sign-off
    No blank pages. No personal names. No redundant sections.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
        from reportlab.platypus import (
            Paragraph, Spacer, Table, TableStyle,
            HRFlowable, KeepTogether, PageBreak, Flowable
        )
        from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate
        from reportlab.lib.colors import HexColor
        import datetime, os, collections, re

        # ── PALETTE ────────────────────────────────────────
        C_BRAND  = HexColor('#7B1C2E')
        C_GOLD   = HexColor('#C49A1A')
        C_GREEN  = HexColor('#1a7a3f')
        C_AMBER  = HexColor('#b45309')
        C_GREY   = HexColor('#475569')
        C_GREY_LT= HexColor('#f9f6f4')
        C_GREY_MD= HexColor('#e8e0dc')
        C_WHITE  = colors.white
        C_BLACK  = HexColor('#1a0e0b')
        C_STRIPE = HexColor('#fdf9f7')

        W, H = A4
        ML, MR, MT, MB = 1.8*cm, 1.8*cm, 2.4*cm, 2.2*cm
        BW = W - ML - MR

        # ── DATA ───────────────────────────────────────────
        data       = engine.load_all()
        all_rows   = data['inventory']
        hist_items = data['history']

        def bucket(c):
            c = (c or '').strip()
            if c in ('Replaced','Disposed'): return c.lower()
            if 'repair' in c.lower(): return 'repair'
            if c == 'New': return 'new'
            return 'good'

        filtered      = [r for r in all_rows
                         if (not dept_filter or r.get('Department/Location','') == dept_filter)
                         and (not type_filter or r.get('Equipment Type','') == type_filter)]
        active        = [r for r in filtered if bucket(r.get('Condition/Status')) not in ('replaced','disposed')]
        replaced_rows = [r for r in filtered if bucket(r.get('Condition/Status')) == 'replaced']
        disposed_rows = [r for r in filtered if bucket(r.get('Condition/Status')) == 'disposed']

        total    = len(active)
        n_new    = sum(1 for r in active if bucket(r.get('Condition/Status')) == 'new')
        n_good   = sum(1 for r in active if bucket(r.get('Condition/Status')) == 'good')
        n_repair = sum(1 for r in active if bucket(r.get('Condition/Status')) == 'repair')
        n_ok     = n_new + n_good
        pct_ok   = round(n_ok / total * 100) if total else 0

        type_counts = collections.Counter(r.get('Equipment Type','Unknown') or 'Unknown' for r in active)
        dept_counts = collections.Counter(r.get('Department/Location','Unknown') or 'Unknown' for r in active)

        today_d = datetime.date.today()
        cutoff  = today_d - datetime.timedelta(days=90)
        def pdate(s):
            try: return datetime.datetime.strptime(str(s)[:10],'%Y-%m-%d').date()
            except: return None
        # Apply date range: if specified use it, else default to last 90 days
        def in_range(h):
            d = pdate(h.get('date',''))
            if d is None: return False
            if date_from:
                try:
                    if d < datetime.datetime.strptime(date_from,'%Y-%m-%d').date(): return False
                except: pass
            if date_to:
                try:
                    if d > datetime.datetime.strptime(date_to,'%Y-%m-%d').date(): return False
                except: pass
            if not date_from and not date_to:
                if d < cutoff: return False
            return True
        recent = [h for h in hist_items if in_range(h)]

        total_cost = 0.0
        for h in recent:
            m = re.search(r'\[Cost: KES ([0-9,.]+)\]', h.get('note',''))
            if m:
                try: total_cost += float(m.group(1).replace(',',''))
                except: pass

        now_str  = datetime.datetime.now().strftime('%d %B %Y  %H:%M')
        date_str = datetime.datetime.now().strftime('%d %B %Y')
        scope    = dept_filter or type_filter or 'All Assets'
        top_type = type_counts.most_common(1)[0][0] if type_counts else '—'
        top_dept = dept_counts.most_common(1)[0][0] if dept_counts else '—'

        # ── STYLES ─────────────────────────────────────────
        def S(name, **kw): return ParagraphStyle(name, **kw)
        s_h2   = S('H2', fontSize=9.5, fontName='Helvetica-Bold', textColor=C_BRAND,
                   spaceAfter=2, spaceBefore=5, leading=12)
        s_exec = S('Exec', fontSize=8.5, fontName='Helvetica', textColor=C_BLACK,
                   leading=13, alignment=TA_JUSTIFY, spaceAfter=3)
        s_foot = S('Foot', fontSize=7, fontName='Helvetica', textColor=C_GREY,
                   alignment=TA_CENTER)

        # ── PAGE TEMPLATE ──────────────────────────────────
        def make_doc(path):
            class MyDoc(BaseDocTemplate):
                def __init__(self, fn, **kw):
                    super().__init__(fn, **kw)
                    frame = Frame(ML, MB, BW, H - MT - MB,
                                  leftPadding=0, rightPadding=0,
                                  topPadding=0, bottomPadding=0, id='main')
                    self.addPageTemplates([
                        PageTemplate(id='main', frames=[frame], onPage=self._hf)])

                def _hf(self, cv, doc):
                    cv.saveState()
                    # ── Header band ──
                    cv.setFillColor(C_BRAND)
                    cv.rect(0, H - 1.4*cm, W, 1.4*cm, fill=1, stroke=0)
                    cv.setFillColor(C_GOLD)
                    cv.rect(0, H - 1.42*cm, W, 0.04*cm, fill=1, stroke=0)

                    lx = ML
                    if logo_path and os.path.exists(logo_path):
                        try:
                            from reportlab.lib.utils import ImageReader
                            img = ImageReader(logo_path)
                            iw, ih = img.getSize()
                            lh = 1.0*cm
                            lw = lh * iw / ih
                            cv.drawImage(logo_path, lx, H - 1.2*cm,
                                         width=lw, height=lh,
                                         preserveAspectRatio=True, mask='auto')
                            lx += lw + 6
                        except: pass

                    cv.setFillColor(C_WHITE)
                    cv.setFont('Helvetica-Bold', 8)
                    cv.drawString(lx, H - 0.75*cm, 'St. Anne Mission Hospital — ICT Department')
                    cv.setFont('Helvetica', 7)
                    cv.setFillColor(HexColor('#c4b4b8'))
                    cv.drawString(lx, H - 1.15*cm, f'ICT Asset Report  ·  {scope}  ·  Serve With Love')
                    cv.setFillColor(C_WHITE)
                    cv.setFont('Helvetica-Bold', 7)
                    cv.drawRightString(W - MR, H - 0.75*cm, date_str)
                    cv.setFont('Helvetica', 6.5)
                    cv.setFillColor(HexColor('#c4b4b8'))
                    cv.drawRightString(W - MR, H - 1.15*cm, f'Page {doc.page}  ·  CONFIDENTIAL')

                    # ── Footer ──
                    cv.setStrokeColor(C_GREY_MD)
                    cv.setLineWidth(0.5)
                    cv.line(ML, MB - 0.3*cm, W - MR, MB - 0.3*cm)
                    cv.setFillColor(C_GREY)
                    cv.setFont('Helvetica', 6.5)
                    cv.drawString(ML, MB - 0.55*cm,
                        f'St. Anne Mission Hospital  ·  ICT Asset Report  ·  Generated {now_str}')
                    cv.drawRightString(W - MR, MB - 0.55*cm, 'CONFIDENTIAL — ICT USE ONLY')
                    cv.restoreState()

            return MyDoc(path, pagesize=A4,
                         leftMargin=ML, rightMargin=MR,
                         topMargin=MT, bottomMargin=MB,
                         title='ICT Asset Report',
                         author='St. Anne Mission Hospital ICT')

        # ── HELPERS ────────────────────────────────────────
        def sp(h=4): return Spacer(1, h)
        def hr():    return HRFlowable(width='100%', thickness=0.6, color=C_GOLD,
                                       spaceAfter=3, spaceBefore=1)

        def sh(text):
            return [sp(4), Paragraph(text, s_h2), hr()]

        def dtable(headers, rows_data, col_widths, hbg=C_BRAND, fs=7.5):
            _th = ParagraphStyle('th', fontSize=fs-0.5, fontName='Helvetica-Bold',
                                 textColor=C_WHITE, leading=fs+1, alignment=TA_CENTER)
            _td = ParagraphStyle('td', fontSize=fs, fontName='Helvetica',
                                 textColor=C_BLACK, leading=fs+1.5, alignment=TA_LEFT)
            _tc = ParagraphStyle('tc', fontSize=fs, fontName='Helvetica',
                                 textColor=C_BLACK, leading=fs+1.5, alignment=TA_CENTER)

            tdata = [[Paragraph(h, _th) for h in headers]]
            for row in rows_data:
                tdata.append([Paragraph(str(c or '—'), _tc if i == 0 else _td)
                               for i, c in enumerate(row)])

            t = Table(tdata, colWidths=col_widths, repeatRows=1)
            style = [
                ('BACKGROUND',    (0,0), (-1,0), hbg),
                ('TOPPADDING',    (0,0), (-1,-1), 2),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
                ('LEFTPADDING',   (0,0), (-1,-1), 4),
                ('RIGHTPADDING',  (0,0), (-1,-1), 4),
                ('GRID',          (0,0), (-1,-1), 0.3, C_GREY_MD),
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('ROWBACKGROUNDS',(0,1), (-1,-1), [C_WHITE, C_STRIPE]),
            ]
            if rows_data and str(rows_data[-1][0]).upper().startswith('TOTAL'):
                style += [
                    ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
                    ('BACKGROUND', (0,-1), (-1,-1), C_GREY_LT),
                    ('LINEABOVE',  (0,-1), (-1,-1), 0.6, C_GREY),
                ]
            t.setStyle(TableStyle(style))
            return t

        # ── KPI STRIP ──────────────────────────────────────
        def kpi_strip():
            items = [
                ('Total Active Assets', str(total),            '#7B1C2E'),
                ('Good / New',          f'{n_ok}  ({pct_ok}%)','#1a7a3f'),
                ('Need Attention',      str(n_repair),          '#b45309'),
                ('Replaced / Retired',  str(len(replaced_rows)+len(disposed_rows)), '#C49A1A'),
                ('Departments',         str(len(dept_counts)),  '#475569'),
            ]
            n   = len(items)
            cw  = BW / n
            vals = [Paragraph(
                        f'<font size="18"><b><font color="{c}">{v}</font></b></font>',
                        ParagraphStyle('kv', alignment=TA_CENTER, leading=22))
                    for _, v, c in items]
            lbls = [Paragraph(
                        f'<font color="#475569"><font size="6.5">{l}</font></font>',
                        ParagraphStyle('kl', alignment=TA_CENTER, leading=9))
                    for l, _, _ in items]
            t = Table([vals, lbls], colWidths=[cw]*n, rowHeights=[26, 13])
            t.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (-1,-1), C_GREY_LT),
                ('BOX',           (0,0), (-1,-1), 0.4, C_GREY_MD),
                ('INNERGRID',     (0,0), (-1,-1), 0.4, C_GREY_MD),
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('TOPPADDING',    (0,0), (-1,0),  4),
                ('BOTTOMPADDING', (0,1), (-1,1),  4),
            ]))
            return t

        # ════════════════════════════════════════════════
        # BUILD STORY
        # ════════════════════════════════════════════════
        story = []

        # ── KPI strip ──────────────────────────────────
        story.append(kpi_strip())
        story.append(sp(6))

        # ── Executive Summary ───────────────────────────
        story += sh('Executive Summary')
        repair_note = (f'{n_repair} asset{"s" if n_repair!=1 else ""} currently require maintenance.' if n_repair
                       else 'No assets are currently flagged for repair.')
        cost_note   = (f' Maintenance costs (last 90 days): <b>KES {total_cost:,.0f}</b>.' if total_cost else '')
        lifecycle_note = (f' {len(replaced_rows)} unit{"s" if len(replaced_rows)!=1 else ""} retired under lifecycle management.' if replaced_rows else '')
        story.append(Paragraph(
            f'The ICT Department manages <b>{total}</b> active asset{"s" if total!=1 else ""} across '
            f'<b>{len(dept_counts)}</b> department{"s" if len(dept_counts)!=1 else ""}. '
            f'<b>{n_ok}</b> ({pct_ok}%) are in good or new condition. {repair_note}'
            f'{cost_note}{lifecycle_note} '
            f'Leading category: <b>{top_type}</b>. Highest asset count: <b>{top_dept}</b>.',
            s_exec))

        # ── Equipment by Type — full width with row number column ──
        story += sh('Asset Breakdown by Equipment Type')

        # Row number | Equipment Type | Total | Good | Repair | %
        # Full page width; narrow No. col on the left margin
        NO_W  = BW * 0.055   # "#" column
        TY_W  = BW * 0.385   # Equipment Type
        TO_W  = BW * 0.12    # Total
        GD_W  = BW * 0.13    # Good
        RP_W  = BW * 0.13    # Repair
        PC_W  = BW * 0.18    # %
        type_cw_full = [NO_W, TY_W, TO_W, GD_W, RP_W, PC_W]

        type_rows = []
        for idx, (etype, cnt) in enumerate(type_counts.most_common(), 1):
            tr  = [r for r in active if r.get('Equipment Type','') == etype]
            rep = sum(1 for r in tr if bucket(r.get('Condition/Status')) == 'repair')
            ok  = cnt - rep
            type_rows.append([str(idx), etype, str(cnt), str(ok), str(rep),
                               f'{round(cnt/total*100)}%' if total else '0%'])
        type_rows.append(['—', 'TOTAL', str(total), str(n_ok), str(n_repair), '100%'])

        story.append(dtable(
            ['No.', 'Equipment Type', 'Total', 'Good', 'Repair', '%'],
            type_rows, type_cw_full))

        # ── Assets Needing Attention ────────────────────
        repair_assets = [r for r in active if bucket(r.get('Condition/Status')) == 'repair']
        if repair_assets:
            story += sh('Assets Requiring Attention')
            r_rows = [[
                r.get('Equipment Type',''),
                r.get('Brand/Model',''),
                r.get('Serial No.',''),
                r.get('Department/Location',''),
                r.get('Condition/Status',''),
            ] for r in repair_assets]
            cw_rep = [BW*0.18, BW*0.22, BW*0.18, BW*0.26, BW*0.16]
            story.append(dtable(
                ['Type','Brand / Model','No.','Department','Status'],
                r_rows, cw_rep, hbg=HexColor('#7B1C2E')))

        # ── Recent Activity ─────────────────────────────
        if recent:
            story += sh('Recent Activity — Last 90 Days')
            ev_counts = collections.Counter(h.get('event','other') for h in recent)
            ev_rows   = [[ev.title(), str(cnt)] for ev, cnt in ev_counts.most_common()]
            ev_rows.append(['TOTAL', str(len(recent))])
            activity_note = (
                f'{len(recent)} event{"s" if len(recent)!=1 else ""} recorded in the past 90 days'
                f'{f" · KES {total_cost:,.0f} in logged costs" if total_cost else ""}.'
            )
            story.append(Paragraph(activity_note, s_exec))
            story.append(dtable(['Event Type','Count'], ev_rows,
                                 [BW*0.55, BW*0.20]))

        # ── Lifecycle (only if data exists) ─────────────
        if replaced_rows or disposed_rows:
            story += sh('Lifecycle Summary')
            rep_types = collections.Counter(r.get('Equipment Type','Unknown') or 'Unknown' for r in replaced_rows)
            dis_types = collections.Counter(r.get('Equipment Type','Unknown') or 'Unknown' for r in disposed_rows)
            all_et  = sorted(set(list(rep_types)+list(dis_types)))
            lc_rows = [[et, str(rep_types.get(et,0)), str(dis_types.get(et,0))] for et in all_et]
            lc_rows.append(['TOTAL', str(len(replaced_rows)), str(len(disposed_rows))])
            story.append(dtable(['Equipment Type','Replaced','Disposed'],
                                 lc_rows, [BW*0.60, BW*0.20, BW*0.20]))

        # ── Sign-off ─────────────────────────────────────
        story.append(sp(8))
        story.append(HRFlowable(width='100%', thickness=0.5, color=C_GREY_MD,
                                spaceAfter=4, spaceBefore=4))
        story.append(Paragraph(
            f'<font color="#7B1C2E"><b>St. Anne Mission Hospital  ·  ICT Department</b></font>'
            f'  &nbsp;·&nbsp;  <i>Serve With Love</i>'
            f'  &nbsp;·&nbsp;  <font color="#94a3b8">Generated {now_str}  ·  CONFIDENTIAL</font>',
            ParagraphStyle('sf', fontSize=7, fontName='Helvetica',
                           alignment=TA_CENTER, textColor=C_GREY)))

        # ── BUILD ─────────────────────────────────────────
        doc = make_doc(output_path)
        doc.build(story)
        engine.log_session('REPORT', f'PDF report generated: {os.path.basename(output_path)}')
        return {"ok": True, "path": output_path}

    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "trace": traceback.format_exc()}


def main():
    engine = ExcelEngine()
    cfg    = ConfigEngine()
    api    = Api(engine, cfg)

    ui_path = os.path.join(BUNDLE_DIR, 'ui.html')
    if not os.path.exists(ui_path):
        print("ERROR: ui.html not found in", BUNDLE_DIR)
        sys.exit(1)

    if "--desktop" in sys.argv:
        run_desktop(api, ui_path)
    elif "--electron" in sys.argv:
        run_browser(api, ui_path, electron_mode=True)
    else:
        run_browser(api, ui_path)

if __name__ == "__main__":
    main()
